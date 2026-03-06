from __future__ import annotations

from pathlib import Path
from werkzeug.security import generate_password_hash
from openpyxl import load_workbook

from models import db, Region, Artist, Song, Tag, Genre, AdminUser


XLSX_PATH = Path(__file__).parent / "Sonidos de mi tierra.xlsx"


def _norm(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v if v != "" else None
    return v


def seed_all(app, admin_username: str, admin_password: str):
    with app.app_context():
        db.create_all()

        # 1) Admin user (idempotente)
        admin = AdminUser.query.filter_by(username=admin_username).first()
        if not admin:
            admin = AdminUser(
                username=admin_username,
                password_hash=generate_password_hash(admin_password),
            )
            db.session.add(admin)
            db.session.commit()

        # 2) Cargar Excel
        if not XLSX_PATH.exists():
            raise FileNotFoundError(f"No encuentro el archivo Excel en: {XLSX_PATH}")

        wb = load_workbook(XLSX_PATH)

        def sheet_by_alias(wb, desired: str):
            # match exact o por minúsculas y sin espacios
            norm = lambda s: s.strip().lower().replace(" ", "_")
            desired_n = norm(desired)

            for s in wb.sheetnames:
                if norm(s) == desired_n:
                    return s

            # alias comunes
            aliases = {
                "artists": ["artist", "artistas"],
                "songs": ["song", "canciones"],
                "regions": ["region", "regiones"],
                "genres": ["genre", "genero", "generos", "géneros"],
                "tags": ["tag", "etiquetas"],
                "song_tags": ["songs_tags", "songtag", "canciones_tags"],
                "song_genres": ["songs_genres", "songgenre", "canciones_genres"],
                "song_regions": ["songs_regions", "songregion", "canciones_regions"],
            }

            for a in aliases.get(desired, []):
                for s in wb.sheetnames:
                    if norm(s) == norm(a):
                        return s

            raise KeyError(f"No existe hoja para '{desired}'. Hojas: {wb.sheetnames}")

        def read_sheet(name: str):
            ws = wb[sheet_by_alias(wb, name)]
            rows = list(ws.iter_rows(values_only=True))

            # busca en las primeras 15 filas la primera que parezca header real
            header_row_idx = None
            for i, r in enumerate(rows[:15]):
                non_empty = [c for c in r if c is not None and str(c).strip() != ""]
                if len(non_empty) >= 2:
                    header_row_idx = i
                    break

            if header_row_idx is None:
                return []

            header = [str(h).strip() if h is not None else "" for h in rows[header_row_idx]]
            data = []

            for r in rows[header_row_idx + 1:]:
                if all(c is None or (isinstance(c, str) and c.strip() == "") for c in r):
                    continue

                row_dict = {}
                for j, key in enumerate(header):
                    if not key:
                        continue
                    row_dict[key] = _norm(r[j] if j < len(r) else None)
                data.append(row_dict)

            return data

        # 3) Importar catálogos primero: regions, genres, tags, artists
        regions_rows = read_sheet("regions")
        for row in regions_rows:
            rid = row.get("id")
            if not rid:
                continue
            obj = Region.query.get(rid)
            if not obj:
                obj = Region(
                    id=rid,
                    name=row.get("name"),
                    parent=row.get("parent"),
                    description=row.get("description"),
                )
                db.session.add(obj)
        db.session.commit()

        genres_rows = read_sheet("genres")
        for row in genres_rows:
            gid = row.get("genre_id")
            if not gid:
                continue
            obj = Genre.query.get(gid)
            if not obj:
                obj = Genre(
                    genre_id=gid,
                    name=row.get("name"),
                    description=row.get("description"),
                )
                db.session.add(obj)
        db.session.commit()

        tags_rows = read_sheet("tags")
        for row in tags_rows:
            tid = row.get("id")
            if not tid:
                continue
            # OJO: Tag.id es int en la mayoría de modelos; si el tuyo es string, avísame.
            obj = Tag.query.get(int(tid))
            if not obj:
                obj = Tag(
                    id=int(tid),
                    name=row.get("name"),
                    category=row.get("category"),
                    description=row.get("description"),
                )
                db.session.add(obj)
        db.session.commit()

        artists_rows = read_sheet("artists")
        print("artists_rows:", len(artists_rows))
        print("artists_header_keys:", list(artists_rows[0].keys()) if artists_rows else "EMPTY")
        for row in artists_rows:
            aid = row.get("id")
            if not aid:
                continue
            obj = Artist.query.get(int(aid))
            if not obj:
                obj = Artist(
                    id=int(aid),
                    name=row.get("name"),
                    photo_url=row.get("photo_url"),
                    bio=row.get("bio"),
                    birth_year=row.get("birth_year"),
                    death_year=row.get("death_year"),
                    origin_region=row.get("origin_region"),
                    era=row.get("era"),
                    cultural_contribuitions=row.get("cultural_contribuitions"),
                )
                db.session.add(obj)
        db.session.commit()

        # 4) Importar songs
        songs_rows = read_sheet("songs")
        print("songs_rows:", len(songs_rows))
        print("songs_header_keys:", list(songs_rows[0].keys()) if songs_rows else "EMPTY")
        for row in songs_rows:
            sid = row.get("id")
            if not sid:
                continue
            obj = Song.query.get(int(sid))
            if not obj:
                obj = Song(
                    id=int(sid),
                    title=row.get("title"),
                    artist_id=int(row.get("artist_id")) if row.get("artist_id") else None,
                    year=int(row.get("year")) if row.get("year") else None,
                    iconic_score=int(row.get("Iconic Score (1-100)") or row.get("iconic_score") or 50),
                    decade=int(row.get("decade")) if row.get("decade") else None,
                    link=row.get("Link") or row.get("link"),
                    description=row.get("Description") or row.get("description"),
                )
                db.session.add(obj)
        db.session.commit()

        # 5) Importar relaciones (tablas puente) usando las relaciones ORM
        # song_tags
        st_rows = read_sheet("song_tags")
        for row in st_rows:
            sid = row.get("song_id")
            tid = row.get("tag_id")
            if not sid or not tid:
                continue
            song = Song.query.get(int(sid))
            tag = Tag.query.get(int(tid))
            if song and tag and tag not in song.tags:
                song.tags.append(tag)
        db.session.commit()

        # song_genres
        sg_rows = read_sheet("song_genres")
        for row in sg_rows:
            sid = row.get("song_id")
            gid = row.get("genre_id")
            if not sid or not gid:
                continue
            song = Song.query.get(int(sid))
            genre = Genre.query.get(str(gid))
            if song and genre and genre not in song.genres:
                song.genres.append(genre)
        db.session.commit()

        # song_regions
        sr_rows = read_sheet("song_regions")
        for row in sr_rows:
            sid = row.get("song_id")
            rid = row.get("region_id")
            if not sid or not rid:
                continue
            song = Song.query.get(int(sid))
            region = Region.query.get(str(rid))
            if song and region and region not in song.regions:
                song.regions.append(region)
        db.session.commit()

        print("✅ Seed completado desde Excel.")